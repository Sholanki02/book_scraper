import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os

# Function to fetch the latest conversion rate from EUR to INR using an API
def get_conversion_rate():
    api_url = "https://api.exchangerate-api.com/v4/latest/EUR"  # Example API URL
    try:
        print("Fetching conversion rate...")
        response = requests.get(api_url)
        response.raise_for_status()
        data = response.json()
        conversion_rate = data['rates']['INR']
        print(f"Conversion rate fetched: {conversion_rate}")
        return conversion_rate
    except requests.exceptions.RequestException as e:
        print(f"Error fetching conversion rate: {e}")
        return 90  # Fallback to a default value

# Function to send an email notification with new books
def send_email(new_books):
    sender_email = "sholanki.brightchamps@gmail.com"
    receiver_email = "samemmanuels01@gmail.com"
    password = "rclu hczc veuf aqrb"

    subject = "New Books Added"

    # Limit to first 10 books
    new_books_limited = new_books[:10]

    # HTML table structure for the email
    body = """
    <html>
    <body>
    <h2>The following new books were added:</h2>
    <table border="1" cellpadding="10" cellspacing="0" style="border-collapse: collapse;">
        <tr>
            <th>Title</th>
            <th>Price (Euro)</th>
            <th>Price (INR)</th>
            <th>Availability</th>
            <th>Rating</th>
        </tr>
    """
    
    # Add rows for each book in the table
    for book in new_books_limited:
        body += f"""
        <tr>
            <td>{book[0]}</td>
            <td>{book[1]}</td>
            <td>{book[2]}</td>
            <td>{book[3]}</td>
            <td>{book[4]}</td>
        </tr>
        """
    
    # Closing the table and body tags
    body += """
    </table>
    </body>
    </html>
    """

    msg = MIMEMultipart()
    msg['From'] = f"Book Store <{sender_email}>"
    msg['To'] = receiver_email 
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'html'))  # Set to 'html' for proper rendering

    try:
        print("Sending email...")
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, password)
            server.send_message(msg)
            print("Email sent successfully.")
    except Exception as e:
        print(f"Failed to send email: {e}")

# Function to save data to Excel, appending if the file exists
def save_to_excel(data, file_name):
    print(f"Saving data to '{file_name}'...")
    if os.path.exists(file_name):
        # Open existing file and append data
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        # Create a new file
        wb = Workbook()
        ws = wb.active
        ws.title = "Books Data"
        # Add headers for a new file
        headers = ['Title', 'Price (Euro)', 'Price (INR)', 'Availability', 'Rating']
        ws.append(headers)

    # Append new data
    for book in data:
        ws.append(book)

    # Adjust column widths
    column_widths = [50, 15, 15, 20, 15]
    for col, width in zip(ws.columns, column_widths):
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = width
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(file_name)
    print(f"Data saved to '{file_name}'.")

# Function to scrape books data from the website
def scrape_books():
    print("Starting the scraping process...")
    base_url = 'http://books.toscrape.com/catalogue/page-{}.html'
    page = 1
    all_books = []
    new_books = []
    euro_to_inr = get_conversion_rate()  # Get live conversion rate

    while True:
        url = base_url.format(page)
        print(f"Scraping page {page}...")
        try:
            response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
            response.raise_for_status()

            # Break if page doesn't exist
            if response.status_code == 404:
                print("No more pages to scrape. Exiting...")
                break

            soup = BeautifulSoup(response.text, 'html.parser')
            books = soup.find_all('article', class_='product_pod')

            if not books:
                break  # No more books on the next page

            for book in books:
                title = book.h3.a['title']
                price_text = book.find('p', class_='price_color').text.strip()
                
                # Clean up unexpected characters in the price string
                price_text = price_text.encode('ascii', 'ignore').decode()
                
                # Extract price and convert to INR
                price_euro = float(price_text.replace('€', '').replace('£', '').strip())
                price_inr = price_euro * euro_to_inr  # Convert to INR
                availability = book.find('p', class_='instock availability').text.strip()

                # Scrape book rating (stars)
                rating_class = book.find('p', class_='star-rating')
                rating = rating_class['class'][1] if rating_class else 'No rating'

                # Append book details (including rating) to the list
                all_books.append([title, f"€{price_euro:.2f}", f"₹{price_inr:.2f}", availability, rating])

                new_books.append([title, f"€{price_euro:.2f}", f"₹{price_inr:.2f}", availability, rating])

            page += 1

        except requests.exceptions.RequestException as e:
            print(f"Error while fetching the page: {e}")
            break
        except ValueError as ve:
            print(f"Error processing price: {ve}")
            break

    # Send email if new books were added
    if new_books:
        send_email(new_books)

    # Save data to Excel (append data)
    save_to_excel(all_books, 'books_with_prices.xlsx')

    print("Scraping process completed.")

if __name__ == "__main__":
    try:
        scrape_books()
    except Exception as e:
        print(f"Unexpected error: {e}")
